import openpyxl
import csv

wb = openpyxl.load_workbook('data/raw/templates/s1_neg.xlsm', data_only=False)

# Text output
with open('analysis/s1_neg_formulas.txt', 'w', encoding='utf-8') as txt_out:
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        txt_out.write(f"Sheet: {sheetname}\n")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    txt_out.write(f"Cell {cell.coordinate}: {cell.value}\n")

# CSV output
with open('data/processed/s1_neg_formulas.csv', 'w', encoding='utf-8', newline='') as csv_out:
    writer = csv.writer(csv_out)
    writer.writerow(['Sheet', 'Cell', 'Formula'])
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    writer.writerow([sheetname, cell.coordinate, cell.value])
