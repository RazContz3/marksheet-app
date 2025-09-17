import openpyxl

wb = openpyxl.load_workbook('data/raw/marksheets/10S1_Marks_Sheet_2025.xlsx', data_only=True)

with open('analysis/10S1_Marks_Sheet_2025_structure.txt', 'w', encoding='utf-8') as out:
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        out.write(f"Sheet: {sheetname}\n")
        # Print first 10 rows and columns
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=15):
            values = [str(cell.value) if cell.value is not None else '' for cell in row]
            out.write("\t".join(values) + "\n")
        out.write("---\n")
